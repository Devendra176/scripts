import argparse
import os
import dotenv
import copy
import openpyxl
import smtplib
import re
import time
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from concurrent.futures import ThreadPoolExecutor, as_completed

from global_var import SUBJECT, BODY, ATTACHMENT, EMAIL_FILE


class SendMail:
    def __init__(self, delay=2.0):
        self.sender_email = sender_email
        self.password = password
        self.host = host
        self.port = int(port)
        self.emails = []
        self.failed_emails = []
        self.sheet = None
        self.delay = delay

    def read_file(self):
        workbook = openpyxl.load_workbook(EMAIL_FILE)
        self.sheet = workbook.active

    def extract_emails(self):
        valid_emails = []
        for row in self.sheet.iter_rows(values_only=True):
            recipient_email = row[0]
            if recipient_email:
                pattern = r"^(?!.*(?:deqode\.com|cis\.com)).*$"
                if re.match(pattern, recipient_email):
                    valid_emails.append(recipient_email.strip())
        return valid_emails

    def send_email(self, recipient_email):
        message = SendMail.get_message_attr(recipient_email)
        try:
            with smtplib.SMTP(self.host, self.port) as server:
                server.starttls()
                server.login(self.sender_email, self.password)
                server.sendmail(self.sender_email, recipient_email, message.as_string())
            print(f"[✓] Email sent to {recipient_email}")
        except Exception as e:
            print(f"[x] Failed to send email to {recipient_email}. Error: {str(e)}")
            self.failed_emails.append((recipient_email, str(e)))
        finally:
            time.sleep(self.delay)

    @staticmethod
    def get_message_attr(recipient_email):
        attachment_resume_obj = AttachResume().clone()
        attachment_resume_obj.set_recipient_email(recipient_email)
        return attachment_resume_obj.message

    def save_failed_emails(self, filename="failed_emails.xlsx"):
        if not self.failed_emails:
            print("\nNo failed emails to save.")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Failed Emails"
        ws.append(["Email", "Error Message"])

        for email, error in self.failed_emails:
            ws.append([email, error])

        wb.save(filename)
        print(f"\n[x] Failed emails saved to {filename}")

    def run(self, max_workers=2):
        self.read_file()
        self.emails = self.extract_emails()

        print(f"Sending {len(self.emails)} emails using {max_workers} threads (delay={self.delay}s)...\n")
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = [executor.submit(self.send_email, email) for email in self.emails]

            for future in as_completed(futures):
                try:
                    future.result()
                except Exception as e:
                    print(f"Thread raised an exception: {e}")

        self.save_failed_emails()


class AttachResume:
    __instance = None

    def __init__(self):
        self._sender_email = sender_email
        self._sender_password = password
        self._subject = SUBJECT
        self._body = BODY
        self._attachment_path = ATTACHMENT
        self._recipient_email = None
        self._message = None
        self.get_message_attr()

    def clone(self):
        AttachResume.__instance = copy.copy(self)
        return AttachResume.__instance

    @property
    def message(self):
        return self._message

    def set_recipient_email(self, recipient):
        self._message["To"] = recipient
        return recipient

    def get_message_attr(self):
        message = MIMEMultipart()
        message["From"] = self._sender_email
        message["Subject"] = self._subject
        message.attach(MIMEText(self._body, "plain"))

        with open(self._attachment_path, "rb") as file:
            attach = MIMEApplication(file.read(), _subtype="pdf")
            attach.add_header(
                "Content-Disposition",
                "attachment",
                filename=os.path.basename(self._attachment_path)
            )
            message.attach(attach)

        self._message = message
        return self._message


if __name__ == "__main__":
    dotenv.load_dotenv(".env")

    parser = argparse.ArgumentParser(description="Gmail SMTP Email Sender Script")
    parser.add_argument("--prod", action="store_true", help="Send using Gmail SMTP (App Password)")
    parser.add_argument("--test", action="store_true", help="Send using test mailtrap")

    parser.add_argument("--threads", type=int, default=2, help="Number of threads (2 recommended for Gmail)")
    parser.add_argument("--delay", type=float, default=2.0, help="Delay between each email in seconds")

    args = parser.parse_args()

    sender_email = password = host = port = None

    if args.prod:
        sender_email = os.getenv("SENDER")
        password = os.getenv("PASSWORD")
        host = os.getenv("HOST", "smtp.gmail.com")
        port = os.getenv("PORT", 587)

    elif args.test:
        sender_email = os.getenv("TEST_SENDER")
        password = os.getenv("TEST_PASSWORD")
        host = os.getenv("TEST_HOST")
        port = os.getenv("TEST_PORT")

    else:
        print("Use the --prod flag to send using Gmail SMTP.")
        exit(1)

    sendmail = SendMail(delay=args.delay)
    sendmail.run(max_workers=args.threads)
